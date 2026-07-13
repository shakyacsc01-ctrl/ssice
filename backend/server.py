from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import csv
import uuid
import bcrypt
import jwt as pyjwt
import logging
import random
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form
from fastapi.responses import JSONResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

# ---------- Setup ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="SSICE LMS API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

JWT_ALG = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_id() -> str:
    return str(uuid.uuid4())


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_token(user_id: str, role: str, minutes: int = 60 * 24 * 7) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=minutes),
        "type": "access",
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user


def require_role(*roles: str):
    async def _dep(user: dict = Depends(get_current_user)) -> dict:
        if user.get("role") not in roles:
            raise HTTPException(403, "Forbidden: insufficient role")
        return user
    return _dep


# ---------- Models ----------
Role = Literal["student", "teacher", "admin"]


class RegisterIn(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: Role = "student"
    phone: Optional[str] = None


class LoginIn(BaseModel):
    email: EmailStr
    password: str
    role: Optional[Role] = None


class CourseIn(BaseModel):
    code: str
    name: str
    duration: str
    fees: float
    description: str = ""
    syllabus: List[str] = []
    image: Optional[str] = None


class MCQIn(BaseModel):
    question: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct: Literal["A", "B", "C", "D"]
    explanation: str = ""
    subject: str
    chapter: str = ""
    difficulty: Literal["easy", "medium", "hard"] = "medium"


class ExamIn(BaseModel):
    title: str
    course_code: str
    duration_min: int = 30
    total_questions: int = 20
    negative_marking: float = 0.0
    pass_percent: float = 40.0
    subject: Optional[str] = None


class AttemptSubmit(BaseModel):
    exam_id: str
    answers: dict  # {mcq_id: "A"/"B"/...}


class NoticeIn(BaseModel):
    title: str
    body: str
    type: Literal["exam", "notes", "fee", "result", "general"] = "general"


class FeeIn(BaseModel):
    student_id: str
    course_code: str
    amount: float
    status: Literal["paid", "due"] = "due"
    note: str = ""


class AttendanceIn(BaseModel):
    student_id: str
    course_code: str
    date: str  # YYYY-MM-DD
    present: bool


# ---------- Auth ----------
def _set_cookie(resp: Response, token: str):
    resp.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=60 * 60 * 24 * 7, path="/")


@api.post("/auth/register")
async def register(body: RegisterIn, response: Response):
    email = body.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already registered")
    if body.role == "admin":
        raise HTTPException(403, "Admins cannot self-register")
    user = {
        "id": new_id(),
        "name": body.name,
        "email": email,
        "phone": body.phone or "",
        "role": body.role,
        "password_hash": hash_password(body.password),
        "created_at": now_iso(),
        "avatar": "",
    }
    await db.users.insert_one(user)
    token = create_token(user["id"], user["role"])
    _set_cookie(response, token)
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"user": user, "token": token}


@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "Invalid email or password")
    if body.role and user["role"] != body.role:
        raise HTTPException(403, f"Account is not a {body.role}")
    token = create_token(user["id"], user["role"])
    _set_cookie(response, token)
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"user": user, "token": token}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ---------- Courses ----------
@api.get("/courses")
async def list_courses():
    rows = await db.courses.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    return rows


@api.get("/courses/{code}")
async def get_course(code: str):
    c = await db.courses.find_one({"code": code}, {"_id": 0})
    if not c:
        raise HTTPException(404, "Course not found")
    return c


@api.post("/courses")
async def create_course(body: CourseIn, user: dict = Depends(require_role("admin", "teacher"))):
    c = {**body.model_dump(), "id": new_id(), "created_at": now_iso()}
    if await db.courses.find_one({"code": c["code"]}):
        raise HTTPException(400, "Course code exists")
    await db.courses.insert_one(c)
    c.pop("_id", None)
    return c


@api.delete("/courses/{code}")
async def delete_course(code: str, user: dict = Depends(require_role("admin"))):
    await db.courses.delete_one({"code": code})
    return {"ok": True}


@api.post("/enroll/{code}")
async def enroll(code: str, user: dict = Depends(require_role("student"))):
    course = await db.courses.find_one({"code": code})
    if not course:
        raise HTTPException(404, "Course not found")
    existing = await db.enrollments.find_one({"student_id": user["id"], "course_code": code})
    if existing:
        return {"ok": True, "already": True}
    await db.enrollments.insert_one({
        "id": new_id(), "student_id": user["id"], "course_code": code,
        "enrolled_at": now_iso(), "progress": 0,
    })
    return {"ok": True}


@api.get("/my/courses")
async def my_courses(user: dict = Depends(require_role("student"))):
    enrolls = await db.enrollments.find({"student_id": user["id"]}, {"_id": 0}).to_list(200)
    codes = [e["course_code"] for e in enrolls]
    courses = await db.courses.find({"code": {"$in": codes}}, {"_id": 0}).to_list(200)
    cmap = {c["code"]: c for c in courses}
    return [{**cmap.get(e["course_code"], {}), "progress": e.get("progress", 0)} for e in enrolls if e["course_code"] in cmap]


# ---------- MCQs ----------
@api.get("/mcqs")
async def list_mcqs(subject: Optional[str] = None, chapter: Optional[str] = None, difficulty: Optional[str] = None, limit: int = 20, random_mode: bool = False):
    q = {}
    if subject: q["subject"] = subject
    if chapter: q["chapter"] = chapter
    if difficulty: q["difficulty"] = difficulty
    if random_mode:
        pipeline = [{"$match": q}, {"$sample": {"size": limit}}, {"$project": {"_id": 0}}]
        return await db.mcqs.aggregate(pipeline).to_list(limit)
    return await db.mcqs.find(q, {"_id": 0}).limit(limit).to_list(limit)


@api.get("/mcqs/subjects")
async def mcq_subjects():
    subs = await db.mcqs.distinct("subject")
    return subs


@api.post("/mcqs")
async def add_mcq(body: MCQIn, user: dict = Depends(require_role("admin", "teacher"))):
    m = {**body.model_dump(), "id": new_id(), "created_at": now_iso()}
    await db.mcqs.insert_one(m)
    m.pop("_id", None)
    return m


@api.delete("/mcqs/{mcq_id}")
async def del_mcq(mcq_id: str, user: dict = Depends(require_role("admin", "teacher"))):
    await db.mcqs.delete_one({"id": mcq_id})
    return {"ok": True}


@api.post("/mcqs/bulk-upload")
async def bulk_upload(file: UploadFile = File(...), user: dict = Depends(require_role("admin", "teacher"))):
    """Accepts CSV with columns: question, option_a, option_b, option_c, option_d, correct, explanation, subject, chapter, difficulty"""
    content = await file.read()
    inserted = 0
    errors = []
    try:
        if file.filename.lower().endswith((".xlsx", ".xls")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]
            data_rows = [dict(zip(headers, r)) for r in rows[1:]]
        else:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            data_rows = [{k.lower().strip(): v for k, v in row.items()} for row in reader]

        for i, row in enumerate(data_rows, start=2):
            try:
                mcq = {
                    "id": new_id(),
                    "question": str(row.get("question", "")).strip(),
                    "option_a": str(row.get("option_a", "")).strip(),
                    "option_b": str(row.get("option_b", "")).strip(),
                    "option_c": str(row.get("option_c", "")).strip(),
                    "option_d": str(row.get("option_d", "")).strip(),
                    "correct": str(row.get("correct", "A")).strip().upper()[:1],
                    "explanation": str(row.get("explanation", "") or "").strip(),
                    "subject": str(row.get("subject", "General")).strip(),
                    "chapter": str(row.get("chapter", "") or "").strip(),
                    "difficulty": str(row.get("difficulty", "medium") or "medium").lower().strip(),
                    "created_at": now_iso(),
                }
                if not mcq["question"] or mcq["correct"] not in ("A", "B", "C", "D"):
                    errors.append(f"Row {i}: invalid")
                    continue
                await db.mcqs.insert_one(mcq)
                inserted += 1
            except Exception as e:
                errors.append(f"Row {i}: {e}")
    except Exception as e:
        raise HTTPException(400, f"Parse failed: {e}")
    return {"inserted": inserted, "errors": errors[:10]}


# ---------- Exams ----------
@api.get("/exams")
async def list_exams():
    return await db.exams.find({}, {"_id": 0}).to_list(200)


@api.get("/exams/{exam_id}")
async def get_exam(exam_id: str, user: dict = Depends(get_current_user)):
    exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam:
        raise HTTPException(404, "Exam not found")
    q = {}
    if exam.get("subject"):
        q["subject"] = exam["subject"]
    pipeline = [{"$match": q}, {"$sample": {"size": exam["total_questions"]}},
                {"$project": {"_id": 0, "correct": 0, "explanation": 0}}]
    questions = await db.mcqs.aggregate(pipeline).to_list(exam["total_questions"])
    return {**exam, "questions": questions}


@api.post("/exams")
async def create_exam(body: ExamIn, user: dict = Depends(require_role("admin", "teacher"))):
    e = {**body.model_dump(), "id": new_id(), "created_at": now_iso()}
    await db.exams.insert_one(e)
    e.pop("_id", None)
    return e


@api.post("/exams/submit")
async def submit_exam(body: AttemptSubmit, user: dict = Depends(require_role("student"))):
    exam = await db.exams.find_one({"id": body.exam_id})
    if not exam:
        raise HTTPException(404, "Exam not found")
    mcq_ids = list(body.answers.keys())
    mcqs = await db.mcqs.find({"id": {"$in": mcq_ids}}, {"_id": 0}).to_list(1000)
    correct = 0
    wrong = 0
    review = []
    for m in mcqs:
        ua = (body.answers.get(m["id"]) or "").upper()
        is_correct = ua == m["correct"]
        if ua and is_correct:
            correct += 1
        elif ua:
            wrong += 1
        review.append({
            "mcq_id": m["id"], "question": m["question"],
            "your_answer": ua, "correct": m["correct"],
            "explanation": m.get("explanation", ""), "is_correct": is_correct,
        })
    total = len(mcqs) or 1
    score = correct - (wrong * exam.get("negative_marking", 0))
    percent = max(0, round((score / total) * 100, 2))
    passed = percent >= exam.get("pass_percent", 40)
    attempt = {
        "id": new_id(), "exam_id": body.exam_id, "exam_title": exam["title"],
        "student_id": user["id"], "student_name": user["name"],
        "course_code": exam.get("course_code"),
        "total": total, "correct": correct, "wrong": wrong,
        "score": round(score, 2), "percent": percent, "passed": passed,
        "submitted_at": now_iso(),
    }
    await db.attempts.insert_one(attempt)

    # auto-create certificate if passed
    if passed:
        existing = await db.certificates.find_one({"student_id": user["id"], "course_code": exam.get("course_code")})
        if not existing and exam.get("course_code"):
            course = await db.courses.find_one({"code": exam["course_code"]})
            cert = {
                "id": new_id(),
                "cert_no": f"SSICE-{datetime.now().year}-{random.randint(10000, 99999)}",
                "student_id": user["id"],
                "student_name": user["name"],
                "course_code": exam["course_code"],
                "course_name": course["name"] if course else exam["course_code"],
                "percent": percent,
                "grade": "A+" if percent >= 90 else "A" if percent >= 75 else "B" if percent >= 60 else "C",
                "issued_at": now_iso(),
                "director": "Pappu Singh",
            }
            await db.certificates.insert_one(cert)
    attempt.pop("_id", None)
    return {"attempt": attempt, "review": review}


# ---------- Results ----------
@api.get("/my/results")
async def my_results(user: dict = Depends(require_role("student"))):
    return await db.attempts.find({"student_id": user["id"]}, {"_id": 0}).sort("submitted_at", -1).to_list(100)


@api.get("/exams/{exam_id}/leaderboard")
async def leaderboard(exam_id: str):
    pipeline = [
        {"$match": {"exam_id": exam_id}},
        {"$sort": {"percent": -1, "submitted_at": 1}},
        {"$limit": 20},
        {"$project": {"_id": 0, "student_name": 1, "percent": 1, "score": 1, "submitted_at": 1}},
    ]
    return await db.attempts.aggregate(pipeline).to_list(20)


# ---------- Certificates ----------
@api.get("/my/certificates")
async def my_certs(user: dict = Depends(require_role("student"))):
    return await db.certificates.find({"student_id": user["id"]}, {"_id": 0}).to_list(100)


@api.get("/certificates/verify/{cert_no}")
async def verify_cert(cert_no: str):
    c = await db.certificates.find_one({"cert_no": cert_no}, {"_id": 0})
    if not c:
        raise HTTPException(404, "Certificate not found")
    return c


# ---------- Notices ----------
@api.get("/notices")
async def list_notices():
    return await db.notices.find({}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)


@api.post("/notices")
async def add_notice(body: NoticeIn, user: dict = Depends(require_role("admin", "teacher"))):
    n = {**body.model_dump(), "id": new_id(), "created_at": now_iso(), "by": user["name"]}
    await db.notices.insert_one(n)
    n.pop("_id", None)
    return n


# ---------- Attendance ----------
@api.post("/attendance")
async def mark_att(body: AttendanceIn, user: dict = Depends(require_role("admin", "teacher"))):
    rec = {**body.model_dump(), "id": new_id(), "marked_by": user["id"]}
    await db.attendance.update_one(
        {"student_id": body.student_id, "course_code": body.course_code, "date": body.date},
        {"$set": rec}, upsert=True,
    )
    return {"ok": True}


@api.get("/my/attendance")
async def my_att(user: dict = Depends(require_role("student"))):
    rows = await db.attendance.find({"student_id": user["id"]}, {"_id": 0}).to_list(500)
    total = len(rows)
    present = sum(1 for r in rows if r["present"])
    return {"rows": rows, "total": total, "present": present, "percent": round((present/total)*100, 1) if total else 0}


# ---------- Fees ----------
@api.post("/fees")
async def add_fee(body: FeeIn, user: dict = Depends(require_role("admin"))):
    f = {**body.model_dump(), "id": new_id(), "created_at": now_iso()}
    await db.fees.insert_one(f)
    f.pop("_id", None)
    return f


@api.get("/my/fees")
async def my_fees(user: dict = Depends(require_role("student"))):
    return await db.fees.find({"student_id": user["id"]}, {"_id": 0}).to_list(100)


@api.get("/fees")
async def all_fees(user: dict = Depends(require_role("admin"))):
    return await db.fees.find({}, {"_id": 0}).to_list(500)


# ---------- Admin ----------
@api.get("/admin/users")
async def all_users(role: Optional[str] = None, user: dict = Depends(require_role("admin"))):
    q = {"role": role} if role else {}
    return await db.users.find(q, {"_id": 0, "password_hash": 0}).to_list(500)


@api.get("/admin/stats")
async def admin_stats(user: dict = Depends(require_role("admin"))):
    return {
        "students": await db.users.count_documents({"role": "student"}),
        "teachers": await db.users.count_documents({"role": "teacher"}),
        "courses": await db.courses.count_documents({}),
        "exams": await db.exams.count_documents({}),
        "mcqs": await db.mcqs.count_documents({}),
        "certificates": await db.certificates.count_documents({}),
        "attempts": await db.attempts.count_documents({}),
    }


# ---------- CORS / startup ----------
app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


SEED_COURSES = [
    ("CCC", "Course on Computer Concepts (CCC)", "80 Hours / 3 Months", 1500, "Govt-recognised entry-level computer literacy program from NIELIT."),
    ("DCA", "Diploma in Computer Application (DCA)", "6 Months", 6000, "Foundation diploma covering MS Office, Internet, Tally basics, and Typing."),
    ("ADCA", "Advanced Diploma in Computer Application", "1 Year", 10000, "Comprehensive advanced diploma covering office automation, accounting, and web basics."),
    ("DCAA", "Diploma in Computer Accounting & Applications", "6 Months", 7000, "Accounting-focused diploma with Tally Prime, GST, and Excel."),
    ("DOAP", "Diploma in Office Automation & Publishing", "6 Months", 6500, "Office Automation, Publishing tools, Email and Internet."),
    ("TALLY", "Tally Prime with GST", "3 Months", 4500, "Tally Prime, GST returns, Inventory, Payroll modules."),
    ("EXCEL", "Advanced Excel", "2 Months", 3500, "Pivot Tables, VLOOKUP, Power Query, Dashboards, Macros."),
    ("WEB", "Web Designing", "3 Months", 5500, "HTML5, CSS3, JavaScript, Bootstrap, Responsive Design."),
    ("HTYP", "Hindi Typing", "2 Months", 1500, "Krutidev / Mangal typing with speed building."),
    ("ETYP", "English Typing", "2 Months", 1500, "Speed and accuracy training for English typing exams."),
    ("BCC", "Basic Computer Course", "2 Months", 1200, "Introduction to computers, Windows, Internet and MS Paint."),
]

SEED_MCQS = [
    ("What does CPU stand for?", "Central Processing Unit", "Computer Personal Unit", "Central Process Unit", "Control Processing Unit", "A", "CPU = Central Processing Unit, the brain of the computer.", "Computer Fundamentals", "Basics", "easy"),
    ("Which of these is an input device?", "Monitor", "Printer", "Keyboard", "Speaker", "C", "Keyboard is an input device.", "Computer Fundamentals", "Hardware", "easy"),
    ("Full form of RAM?", "Random Access Memory", "Read Active Memory", "Ready Access Memory", "Random Active Memory", "A", "RAM = Random Access Memory.", "Computer Fundamentals", "Hardware", "easy"),
    ("Which key combination copies selected text?", "Ctrl+V", "Ctrl+C", "Ctrl+X", "Ctrl+Z", "B", "Ctrl+C copies. Ctrl+V pastes.", "MS Office", "MS Word", "easy"),
    ("In Excel, which function returns the largest value?", "SUM", "AVG", "MAX", "MIN", "C", "MAX returns the largest value.", "MS Office", "MS Excel", "medium"),
    ("Default file extension for Word documents (modern)?", ".doc", ".docx", ".txt", ".rtf", "B", ".docx is the default since Word 2007.", "MS Office", "MS Word", "easy"),
    ("Which is NOT a web browser?", "Chrome", "Firefox", "Linux", "Edge", "C", "Linux is an operating system.", "Internet", "Browsers", "easy"),
    ("HTML stands for?", "Hyper Text Markup Language", "High Text Markup Language", "Hyperlink Text Markup", "Hyper Tabular Markup", "A", "HTML = Hyper Text Markup Language.", "Web", "HTML", "easy"),
    ("Which tag creates a hyperlink in HTML?", "<link>", "<a>", "<href>", "<url>", "B", "Anchor tag <a> creates hyperlinks.", "Web", "HTML", "easy"),
    ("In Tally, GST stands for?", "General Sales Tax", "Goods and Services Tax", "Government Service Tax", "Global Sales Tax", "B", "GST = Goods and Services Tax.", "Tally", "GST", "easy"),
    ("Shortcut to save in Tally?", "Ctrl+A", "Ctrl+S", "Enter", "F12", "C", "Pressing Enter (or Ctrl+A) saves in Tally.", "Tally", "Basics", "medium"),
    ("Which device stores data permanently?", "RAM", "Cache", "Hard Disk", "Register", "C", "Hard disk is non-volatile storage.", "Computer Fundamentals", "Storage", "easy"),
    ("1 KB equals how many bytes?", "100", "1000", "1024", "512", "C", "1 KB = 1024 bytes.", "Computer Fundamentals", "Units", "medium"),
    ("Which protocol is used to send email?", "HTTP", "FTP", "SMTP", "SNMP", "C", "SMTP = Simple Mail Transfer Protocol.", "Internet", "Protocols", "medium"),
    ("CSS is used for?", "Database", "Styling Web Pages", "Server Logic", "Networking", "B", "CSS styles HTML elements.", "Web", "CSS", "easy"),
    ("Which is volatile memory?", "ROM", "Hard Disk", "RAM", "SSD", "C", "RAM loses content when power is off.", "Computer Fundamentals", "Hardware", "medium"),
    ("Spreadsheet software example?", "MS Word", "MS Excel", "MS PowerPoint", "Notepad", "B", "MS Excel is spreadsheet software.", "MS Office", "MS Excel", "easy"),
    ("Which is an operating system?", "MS Word", "Windows 11", "Chrome", "Photoshop", "B", "Windows 11 is an OS.", "Computer Fundamentals", "OS", "easy"),
    ("Pivot table is used in?", "MS Word", "MS Excel", "MS Access", "Notepad", "B", "Pivot tables summarize data in Excel.", "MS Office", "MS Excel", "medium"),
    ("Default port for HTTP?", "21", "25", "80", "443", "C", "HTTP uses port 80.", "Internet", "Protocols", "hard"),
]


async def seed_data():
    # Indexes
    await db.users.create_index("email", unique=True)
    await db.courses.create_index("code", unique=True)
    await db.certificates.create_index("cert_no", unique=True)

    # Admin
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    if not await db.users.find_one({"email": admin_email}):
        await db.users.insert_one({
            "id": new_id(), "name": "Admin", "email": admin_email, "phone": "",
            "role": "admin", "password_hash": hash_password(os.environ["ADMIN_PASSWORD"]),
            "created_at": now_iso(), "avatar": "",
        })
    # Teacher
    t_email = os.environ["TEACHER_EMAIL"].lower()
    if not await db.users.find_one({"email": t_email}):
        await db.users.insert_one({
            "id": new_id(), "name": "Rajesh Kumar", "email": t_email, "phone": "",
            "role": "teacher", "password_hash": hash_password(os.environ["TEACHER_PASSWORD"]),
            "created_at": now_iso(), "avatar": "",
        })
    # Student
    s_email = os.environ["STUDENT_EMAIL"].lower()
    student = await db.users.find_one({"email": s_email})
    if not student:
        student_doc = {
            "id": new_id(), "name": "Amit Sharma", "email": s_email, "phone": "9876543210",
            "role": "student", "password_hash": hash_password(os.environ["STUDENT_PASSWORD"]),
            "created_at": now_iso(), "avatar": "",
        }
        await db.users.insert_one(student_doc)
        student = student_doc

    # Courses
    for code, name, dur, fee, desc in SEED_COURSES:
        if not await db.courses.find_one({"code": code}):
            await db.courses.insert_one({
                "id": new_id(), "code": code, "name": name, "duration": dur, "fees": fee,
                "description": desc, "syllabus": [
                    "Introduction & Basics", "Practical Sessions", "Hands-on Projects",
                    "Mock Tests", "Final Examination",
                ],
                "created_at": now_iso(),
            })

    # MCQs
    if await db.mcqs.count_documents({}) == 0:
        for q, a, b, c, d, corr, exp, subj, ch, diff in SEED_MCQS:
            await db.mcqs.insert_one({
                "id": new_id(), "question": q,
                "option_a": a, "option_b": b, "option_c": c, "option_d": d,
                "correct": corr, "explanation": exp, "subject": subj,
                "chapter": ch, "difficulty": diff, "created_at": now_iso(),
            })

    # Exams
    if await db.exams.count_documents({}) == 0:
        await db.exams.insert_one({
            "id": new_id(), "title": "CCC Mock Test", "course_code": "CCC",
            "duration_min": 15, "total_questions": 10, "negative_marking": 0.0,
            "pass_percent": 40, "subject": None, "created_at": now_iso(),
        })
        await db.exams.insert_one({
            "id": new_id(), "title": "Tally + GST Practice Test", "course_code": "TALLY",
            "duration_min": 10, "total_questions": 5, "negative_marking": 0.25,
            "pass_percent": 50, "subject": "Tally", "created_at": now_iso(),
        })

    # Notices
    if await db.notices.count_documents({}) == 0:
        for t, b, ty in [
            ("New Batch Starting Feb 15", "Admissions open for DCA & ADCA. Limited seats.", "general"),
            ("Final Exam Schedule Released", "CCC final exam will be held on 25th Feb.", "exam"),
            ("Fee Reminder", "Last date to deposit Q1 fee is 28th Feb.", "fee"),
        ]:
            await db.notices.insert_one({
                "id": new_id(), "title": t, "body": b, "type": ty,
                "created_at": now_iso(), "by": "Admin",
            })

    # Enroll student in 2 courses + sample fee
    if student:
        sid = student["id"]
        for code in ["CCC", "TALLY"]:
            if not await db.enrollments.find_one({"student_id": sid, "course_code": code}):
                await db.enrollments.insert_one({
                    "id": new_id(), "student_id": sid, "course_code": code,
                    "enrolled_at": now_iso(), "progress": random.randint(20, 80),
                })
        if not await db.fees.find_one({"student_id": sid}):
            await db.fees.insert_one({
                "id": new_id(), "student_id": sid, "course_code": "CCC",
                "amount": 1500, "status": "paid", "note": "Semester 1",
                "created_at": now_iso(),
            })
            await db.fees.insert_one({
                "id": new_id(), "student_id": sid, "course_code": "TALLY",
                "amount": 4500, "status": "due", "note": "Q1",
                "created_at": now_iso(),
            })

    # test_credentials
    creds_path = Path("/app/memory/test_credentials.md")
    creds_path.parent.mkdir(parents=True, exist_ok=True)
    creds_path.write_text(
        "# SSICE LMS Test Credentials\n\n"
        f"- Admin: {os.environ['ADMIN_EMAIL']} / {os.environ['ADMIN_PASSWORD']} (role: admin)\n"
        f"- Teacher: {os.environ['TEACHER_EMAIL']} / {os.environ['TEACHER_PASSWORD']} (role: teacher)\n"
        f"- Student: {os.environ['STUDENT_EMAIL']} / {os.environ['STUDENT_PASSWORD']} (role: student)\n\n"
        "## Auth endpoints\n"
        "- POST /api/auth/login {email,password,role}\n"
        "- POST /api/auth/register\n"
        "- GET /api/auth/me\n"
    )


@app.on_event("startup")
async def on_start():
    try:
        await seed_data()
        logger.info("Seed data ready.")
    except Exception as e:
        logger.exception(f"Seed failed: {e}")


@app.on_event("shutdown")
async def on_stop():
    client.close()
